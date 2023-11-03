import sys
import os
import openpyxl
import pandas
import xlwings

SOFTWARE_VERSION:str = "v1.0.0"
    
def terminalPrint(_status:bool, _message:str, _printOverlay:bool = False):  #ターミナルにメッセージを表示する   返り値:なし
    if(_status):
        _printStatus = "  OK  "
    else:
        _printStatus = "FAILED"
    
    if(_printOverlay):
        print("[{}]  {}".format(_printStatus, _message), end="\r")
    else:
        print("[{}]  {}".format(_printStatus, _message))

def oscilloscopeFilePath(_directoryPath:str, _dataName:str, _channel:str, _dataType:str): #ALLxxxxからオシロスコープのExcelファイルの絶対パスを生成する 返り値:絶対パス
    return "{}\\{}\\F{}{}.{}".format(_directoryPath, _dataName, _dataName[3:], _channel, _dataType)

def csvToXlsx(_filePath:str):   #csvファイルをxlsxに変換する    返り値:出力ファイルの絶対パス
    try:
        _inputFile = pandas.read_csv(_filePath)
        terminalPrint(True, "read {}".format(_filePath))
    except:
        terminalPrint(False, "E001 : read {}".format(_filePath))
        sys.exit()
    
    try:
        _outputFilePath:str = "{}xlsx".format(_filePath[:-3])
        _inputFile.to_excel(_outputFilePath)
        terminalPrint(True, "convert to excel")
        return _outputFilePath
    except:
        terminalPrint(False, "E002 : convert to excel")
        sys.exit()
    

class _cell:    #高度なセルの操作
    def __init__(self):
        pass
        
    def openSheet(self, _filePath:str, _sheetName:str): #読み取る対象のシートを開く 返り値:なし
        try:
            self._excelFilePath = _filePath
            self._excelFile = xlwings.Book(_filePath)
            self._workSheet = self._excelFile.sheets[_sheetName]
            terminalPrint(True, "open {}".format(_filePath))
        except:
            terminalPrint(False, "E003 : open {}".format(_filePath))
            sys.exit()
        
    def getValue(self, _cellAddreess, _printOverlay:bool = False):  #任意のセルの値を読み取る   返り値:読み取ったセルの値
        _readValue = self._workSheet.range(_cellAddreess).value
        try:
            terminalPrint(True, "read value = {:f}".format(float(_readValue)), _printOverlay)
            return float(_readValue)
        except:
            terminalPrint(True, "read value = {}".format(str(_readValue)))
            return self._workSheet.range(_cellAddreess).value
    
    def closeSheet(self): #読み取る対象のシートを閉じる  返り値:なし
        try:
            self._excelFile.close()
            terminalPrint(True, "close {}".format(self._excelFilePath))
        except:
            terminalPrint(False, "E004 : close {}".format(self._excelFilePath))
            sys.exit()

    
class _dataBase:    #データベースファイルの操作
    def __init__(self, _filePath:str):  #データベースのファイルを指定する   返り値:なし
        self._excelFilePath:str = _filePath
        terminalPrint(True, "specify {} as dataBase file".format(self._excelFilePath))
        
    def __del__(self):  #データベースファイルを閉じる   返り値:なし
        try:
            self._workBook.close()
            terminalPrint(True, "close dataBase file")
        except:
            terminalPrint(False, "E005 : close dataBase file")
            sys.exit()

    def openSheet(self):    #シートを開く   返り値:なし
        try:
            self._workBook = openpyxl.load_workbook(filename=self._excelFilePath, read_only=True)
            self._workSheet = self._workBook["DataBase"]
            terminalPrint(True, "open dataBase sheet")
        except:
            terminalPrint(False, "E006 : open dataBase sheet")
            sys.exit()

    def readCellValue(self, _column:int, _row:int):  #任意のセルの値を読み取る  返り値:なし
        try:
            _readValue = self._workSheet.cell(column=_column, row=_row).value
            terminalPrint(True, "read value = {}".format(_readValue))
            return _readValue
        except:
            terminalPrint(False, "E007 : read value")
            sys.exit()


class _approximateFomula:  #近似式の導出
    def __init__(self): #t=0の行を指定する  返り値:なし
        self._readStartRow:int = 1251
    
    def openSheet(self, _filePath:str, _frequency:int): #対象のシートを開く 返り値:なし
        self._period:float = 1.0 / float(_frequency)
        self._excelFilePath = _filePath
        try:
            self._workBook = openpyxl.load_workbook(self._excelFilePath)
            self._mainWorkSheet = self._workBook["Sheet1"]
            terminalPrint(True, "open sheet")
        except:
            terminalPrint(False, "E008 : open sheet")
            sys.exit()

    def findOneCycle(self): #1周期分の行の範囲を導出    返り値:なし
        _nowColumn = self._readStartRow
        
        while(True):
            if(float(self._mainWorkSheet.cell(column=5, row=_nowColumn).value) > self._period):
                _nowColumn -= 1
                break
            else:
                _nowColumn += 1
        
        terminalPrint(True, "find range of 1 Cycle")
        
        self._readEndRow:int = _nowColumn
        self._toEndRow = None
    
    def extractOneValue(self):  #1周期分の想定データを別のファイルに抽出    返り値:なし
        try:
            self._workBook.create_sheet(title="forCalculation")
            self._calculationSheet = self._workBook["forCalculation"]
            
            terminalPrint(True, "create forCaluculation sheet")
        except:
            terminalPrint(False, "E009 : create forCaluculation sheet")
            sys.exit()
        
        _originalNowColumn:int = 5
        _originalNowRow:int = self._readStartRow
        _toNowColumn:int = 1
        _toNowRow:int = 1
        
        while (True):
            if(_originalNowRow <= self._readEndRow):
                self._calculationSheet.cell(column=_toNowColumn, row=_toNowRow).value = self._mainWorkSheet.cell(column=_originalNowColumn, row=_originalNowRow).value
                
                _originalNowRow += 1
                _toNowRow += 1
            else:
                break
        
        _originalNowColumn += 1
        _originalNowRow = self._readStartRow
        _toNowColumn += 1
        _toNowRow = 1
            
        while (True):
            if(_originalNowRow <= self._readEndRow):
                self._calculationSheet.cell(column=_toNowColumn, row=_toNowRow).value = self._mainWorkSheet.cell(column=_originalNowColumn, row=_originalNowRow).value
                
                _originalNowRow += 1
                _toNowRow += 1
            else:
                break
        
        if((_toNowRow - 2) == (self._readEndRow - self._readStartRow)):
            terminalPrint(True, "copy extract for the relevant value")
            self._toEndRow = _toNowRow - 1
        else:
            terminalPrint(False, "E010 : copy extract For the relevant value")
            sys.exit()
    
    def enterApproximateFomula(self):    #近似式の次数ごとの係数をセルに入力 返り値:なし
        self._calculationSheet["E2"] = "近似式"
        self._calculationSheet["E3"] = "指数"
        self._calculationSheet["F3"] = "近似式の係数"
        
        for _i in range(10, -1, -1):
            self._calculationSheet.cell(column=5, row=(14 -_i)).value = int(_i)
            
            if(_i != 0):
                self._calculationSheet.cell(column=6, row=(14 -_i)).value = "=INDEX(LINEST(B1:B{:d},A1:A{:d}^{{10,9,8,7,6,5,4,3,2,1}}),1,E{:d})".format(self._toEndRow, self._toEndRow, 14 -_i)
            else:
                self._calculationSheet.cell(column=6, row=(14 -_i)).value = "=INDEX(LINEST(B1:B{:d},A1:A{:d}^{{10,9,8,7,6,5,4,3,2,1}}),1,11)".format(self._toEndRow, self._toEndRow)
                
        terminalPrint(True, "enter approximate formula")
        
    def enterApproximateValue(self):    #近似値をセルに入力 返り値:なし
        for _i in range(1, self._toEndRow + 1, 1):
            self._calculationSheet.cell(column=3, row=_i).value = "=F4*(A{}^E4)+F5*(A{}^E5)+F6*(A{}^E6)+F7*(A{}^E7)+F8*(A{}^E8)+F9*(A{}^E9)+F10*(A{}^E10)+F11*(A{}^E11)+F12*(A{}^E12)+F13*(A{}^E13)+F14".format(_i, _i, _i, _i, _i, _i, _i, _i, _i, _i)
        
        terminalPrint(True, "enter approximate value")
    
    def enterMaximumTime(self):  #近似式の最大値をセルに入力 返り値:なし
        self._calculationSheet["E16"] = "yの最大値"
        
        self._calculationSheet["E17"] = "=MAX(C1:C{:d})".format(self._toEndRow)
        
        terminalPrint(True, "enter y-axis max value")
        
    def enterPeakToPeak(self):  #近似値のピーク-ピーク値をセルに入力 返り値:なし
        self._calculationSheet["F16"] = "ピーク-ピーク値"
        
        self._calculationSheet["F17"] = "=ABS(MAX(C1:C{:d}) - MIN(C1:C{:d}))".format(self._toEndRow, self._toEndRow)
    
    def saveApproximateFile(self):  #近似式等を入力したファイルを保存する   返り値:保存したファイルの絶対パス
        try:
            self._workBook.save("{}_TemporaryData.xlsx".format(self._excelFilePath))
            terminalPrint(True, "save approximate file")
            return "{}_TemporaryData.xlsx".format(self._excelFilePath)
        except:
            terminalPrint(False, "E011 : save approximate file")
            sys.exit()
        
    def closeSheet(self):   #対象のシートを閉じる   返り値:なし
        try:
            self._workBook.close()
            terminalPrint(True, "close sheet")
        except:
            terminalPrint(True, "close sheet")
            

class _readEachValue:   #ファイルから各値を読み取る
    def __init__(self):
        pass
        
    def openSheet(self, _filePath:str): #対象のシートを開く 返り値:なし
        self._excelFilePath:str = _filePath
        
        try:
            self._workBook = openpyxl.load_workbook(self._excelFilePath)
            self._workSheet = self._workBook["forCalculation"]
            
            terminalPrint(True, "open sheet")
        except:
            terminalPrint(False, "E012 : open sheet")
            sys.exit()
    
    def findPhasePeak(self):    #ピーク値をとるときの時間を導出 返り値:導出された時間
        cell.openSheet(self._excelFilePath, "forCalculation")
        
        _yMaxValue:float = cell.getValue("E17")
        
        _readNowRow:int = 1
        
        while(True):
            _readNowRowValue = cell.getValue("C{:d}".format(_readNowRow), True)
            if(_readNowRowValue == None):
                terminalPrint(False, "E013 : find peak value")
                self._workBook.close()
                sys.exit()
            elif(_readNowRowValue == _yMaxValue):
                terminalPrint(True, "find peak value                    ")
                _phasePeakValue = self._workSheet.cell(column=1, row=_readNowRow).value
                break
            
            _readNowRow += 1
        
        return _phasePeakValue
    
    def findPeakPeakValue(self):    #ピーク-ピーク値を導出  返り値:ピーク-ピーク値
        return cell.getValue("F17")
    
    def closeSheet(self):   #対象のシートを閉じる   返り値:なし
        try:
            self._workBook.close()
            cell.closeSheet()
            
            terminalPrint(True, "close sheet")
        except:
            terminalPrint(False, "E014 : close sheet")
            sys.exit()


class _main:    #メインプログラム
    def __init__(self):
        pass
    
    def creatFile(self):    #計算結果を格納するファイルの作成   返り値:なし
        try:
            self._newWorkBook = openpyxl.Workbook()
            self._newWorkSheet = self._newWorkBook.active
            terminalPrint(True, "create output file")
        except:
            terminalPrint(False, "E016 : create output file")
            sys.exit()
        
        self._newWorkSheet.title = "計算結果"
        
        self._newWorkSheet["A1"] = "ファイル詳細"
        self._newWorkSheet["E1"] = "位相差"
        self._newWorkSheet["H1"] = "増幅率"
        
        self._newWorkSheet["A2"] = "周波数(Hz)"
        self._newWorkSheet["B2"] = "{}ファイル名".format(READ_DATA1)
        self._newWorkSheet["C2"] = "{}ファイル名".format(READ_DATA2)
        self._newWorkSheet["E2"] = "周波数(Hz)"
        self._newWorkSheet["F2"] = "{}-{}".format(READ_DATA2, READ_DATA1)
        self._newWorkSheet["H2"] = "周波数(Hz)"
        self._newWorkSheet["I2"] = "{}/{}".format(READ_DATA2, READ_DATA1)
        
        try:
            self._newWorkBook.save(OUTPUT_FILE_PATH)
            terminalPrint(True, "temporary save output file")
        except:
            terminalPrint(False, "E017 : temporary save output file")
            sys.exit()
        
    def openSheet(self):    #計算結果を格納するシートを開く 返り値:なし
        try:
            self._workBook = openpyxl.load_workbook(OUTPUT_FILE_PATH)
            self._workSheet = self._workBook["計算結果"]
            terminalPrint(True, "open {}".format(OUTPUT_FILE_PATH))
        except:
            terminalPrint(False, "E018 : open {}".format(OUTPUT_FILE_PATH))
            sys.exit()
        
    def derivationPhaseRatio(self, _row:int):   #位相差及び増幅率を導出&ファイルに保存  返り値:まだ計算していないファイルがあるか
        _frequency = dataBase.readCellValue(1, _row)
        if(_frequency == None):
            return False
        
        terminalPrint(True, "find the phase difference and amplification factor of the {:d} line".format(_row))
        
        _OSCFile1Path:str = csvToXlsx(oscilloscopeFilePath(DATA_DIRECTORY_PATH, dataBase.readCellValue(2, _row), READ_DATA1, "CSV"))
        _OSCFile2Path:str = csvToXlsx(oscilloscopeFilePath(DATA_DIRECTORY_PATH, dataBase.readCellValue(2, _row), READ_DATA2, "CSV"))
        
        approximateFomula.openSheet(_OSCFile1Path, _frequency)
        approximateFomula.findOneCycle()
        approximateFomula.extractOneValue()
        approximateFomula.enterApproximateFomula()
        approximateFomula.enterApproximateValue()
        approximateFomula.enterMaximumTime()
        approximateFomula.enterPeakToPeak()

        approximateFilePath = approximateFomula.saveApproximateFile()
        
        approximateFomula.closeSheet()
        readEachValue.openSheet(approximateFilePath)

        _OSCFile1Phase:float = readEachValue.findPhasePeak()
        _OSCFile1PeakPeak:float = readEachValue.findPeakPeakValue()

        readEachValue.closeSheet()
        
        approximateFomula.openSheet(_OSCFile2Path, _frequency)
        approximateFomula.findOneCycle()
        approximateFomula.extractOneValue()
        approximateFomula.enterApproximateFomula()
        approximateFomula.enterApproximateValue()
        approximateFomula.enterMaximumTime()
        approximateFomula.enterPeakToPeak()
        
        approximateFilePath = approximateFomula.saveApproximateFile()
        
        approximateFomula.closeSheet()
        readEachValue.openSheet(approximateFilePath)

        _OSCFile2Phase:float = readEachValue.findPhasePeak()
        _OSCFile2PeakPeak:float = readEachValue.findPeakPeakValue()
        
        if((float(_OSCFile2Phase) - float(_OSCFile1Phase)) >= 0):
            self._phase:float = 360 * ((float(_OSCFile2Phase) - float(_OSCFile1Phase)) / (1 / _frequency))
        else:
            self._phase:float = 360 * (((1 / _frequency) - abs(float(_OSCFile2Phase) - float(_OSCFile1Phase))) / (1 / _frequency))
        
        self._ratio:float = float(_OSCFile2PeakPeak) / float(_OSCFile1PeakPeak)
        
        terminalPrint(True, "phase contrast = {:f} (°)".format(self._phase))
        terminalPrint(True, "amplification ratio = {:f}".format(self._ratio))

        readEachValue.closeSheet()
        
        self._workSheet.cell(column=1, row=_row + 1).value = _frequency
        self._workSheet.cell(column=2, row=_row + 1).value = _OSCFile1Path[-12:]
        self._workSheet.cell(column=3, row=_row + 1).value = _OSCFile2Path[-12:]
        
        self._workSheet.cell(column=5, row=_row + 1).value = _frequency
        self._workSheet.cell(column=6, row=_row + 1).value = self._phase
        
        self._workSheet.cell(column=8, row=_row + 1).value = _frequency
        self._workSheet.cell(column=9, row=_row + 1).value = self._ratio
        
        terminalPrint(True, "finish enter result : row = {}".format(str(_row)))
        
        return True

    def closeSheet(self):   #計算結果を格納するファイルを閉じる 返り値:なし
        try:
            self._workBook.save(OUTPUT_FILE_PATH)
            terminalPrint(True, "save output file")
        except:
            terminalPrint(False, "E019 : save output file")
            sys.exit()


print("*** Start Ditel Easy Excel Phase Contrast Program ***");
terminalPrint(True, "version : {}".format(SOFTWARE_VERSION))

try:
    READ_DATA1:str = sys.argv[1]
    READ_DATA2:str = sys.argv[2]
    DATA_DIRECTORY_PATH:str = sys.argv[3]
    OUTPUT_FILE_PATH:str = "{}\\{}.xlsx".format(DATA_DIRECTORY_PATH, sys.argv[4])
    
    if((READ_DATA1 != "CH1") and (READ_DATA1 != "CH2") and (READ_DATA1 != "MTH")):
        raise ValueError(None)
    
    if((READ_DATA2 != "CH1") and (READ_DATA2 != "CH2") and (READ_DATA2 != "MTH")):
        raise ValueError(None)
    
    terminalPrint(True, "read data information")
    terminalPrint(True, "read data directory Path = {}".format(DATA_DIRECTORY_PATH))
    terminalPrint(True, "read Data1 type = {}".format(READ_DATA1))
    terminalPrint(True, "read Data2 type = {}".format(READ_DATA2))
    terminalPrint(True, "output file name = {}".format(OUTPUT_FILE_PATH))
except:
    terminalPrint(False, "E015 : read data Information")
    print("please input data directory path")
    print('Ex : python3 ./Ditel_Easy_Excel_Phase "Data1 Type" "Data2 Type" "data directory path" "output file name"')
    print('only "CH1", "CH2" or "MTH" can be entered for "DATA Type" and "DATA2 Type"')
    sys.exit()
    
DATA_BASE_FILE_PAHT:str = "{}\\{}".format(os.getcwd(), "dataBase.xlsx")

cell = _cell()
dataBase = _dataBase(DATA_BASE_FILE_PAHT)
approximateFomula = _approximateFomula()
readEachValue = _readEachValue()
main = _main()

dataBase.openSheet()

main.creatFile()
main.openSheet()

_nowRow:int = 2

while (True):
    if(main.derivationPhaseRatio(_nowRow) == False):
        break
    else:
        _nowRow += 1

terminalPrint(True, "finish all derivation : end row = {:d}".format(_nowRow))
main.closeSheet()

terminalPrint(True, "finish all program")