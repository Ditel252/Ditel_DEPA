import sys
import os
import openpyxl
import pandas
import xlwings

#定数
SOFTWARE_VERSION:str = "v0.1.0"
    
def terminalPrint(_status:bool, _message:str):  #ターミナルにメッセージを表示する   返り値:なし
    if(_status):
        _printStatus = "  OK  "
    else:
        _printStatus = "FAILED"
    
    print("[{}]  {}".format(_printStatus, _message))

def oscilloscopeFilePath(_directoryPath:str, _dataName:str, _channel:str, _dataType:str): #ALLxxxxからオシロスコープのExcelファイルの絶対パスを生成する 返り値:絶対パス
    return "{}\\{}\\F{}{}.{}".format(_directoryPath, _dataName, _dataName[3:], _channel, _dataType)

def csvToXlsx(_filePath:str):   #csvファイルをxlsxに変換する    返り値:出力ファイルの絶対パス
    try:
        _inputFile = pandas.read_csv(_filePath)
        terminalPrint(True, "read {}".format(_filePath))
    except:
        terminalPrint(False, "read {}".format(_filePath))
        exit()
    
    try:
        _outputFilePath:str = "{}xlsx".format(_inputFile[:-3])
        _inputFile.to_excel(_outputFilePath)
        terminalPrint(True, "convert to excel")
        return _outputFilePath
    except:
        terminalPrint(False, "convert to excel")
        exit()
    

class _cell:    #高度なセルの操作
    def __init__(self):
        pass
        
    def openSheet(self, _filePath:str, _sheetName:str): #読み取る対象のシートを開く 返り値:なし
        try:
            self._excelFilePath = _filePath
            self._excelFile = xlwings.Book(_filePath)
            self._workSheet = self._excelFile.sheets[_sheetName]
            terminalPrint(True, "open {}".format(_filePath))
        except:
            terminalPrint(False, "open {}".format(_filePath))
            exit()
        
    def getValue(self, _cellAddreess):  #任意のセルの値を読み取る   返り値:読み取ったセルの値
        _readValue = self._workSheet.range(_cellAddreess).value
        try:
            terminalPrint(True, "read value = {:f}".format(float(_readValue)))
            return float(_readValue)
        except:
            terminalPrint(True, "read value = {}".format(str(_readValue)))
            return self._workSheet.range(_cellAddreess).value
    
    def closeSheet(self): #読み取る対象のシートを閉じる  返り値:なし
        try:
            self._excelFile.close()
            terminalPrint(True, "close {}".format(self._excelFilePath))
        except:
            terminalPrint(False, "close {}".format(self._excelFilePath))
            exit()

    
class _dataBase:    #データベースファイルの操作
    def __init__(self, _filePath:str):  #データベースのファイルを指定する   返り値:なし
        self._excelFilePath:str = _filePath
        terminalPrint(True, "specify {} as dataBase file".format(self._excelFilePath))
        
    def __del__(self):  #データベースファイルを閉じる   返り値:なし
        try:
            self._workBook.close()
            terminalPrint(True, "close dataBase file")
        except:
            terminalPrint(False, "close dataBase file")
            exit()

    def openSheet(self):    #シートを開く   返り値:なし
        try:
            self._workBook = openpyxl.load_workbook(filename=self._excelFilePath, read_only=True)
            self._workSheet = self._workBook["DataBase"]
            terminalPrint(True, "open dataBase sheet")
        except:
            terminalPrint(False, "open dataBase sheet")
            exit()

    def readCellValue(self, _column:int, _row:int):  #任意のセルの値を読み取る  返り値:なし
        try:
            _readValue = self._workSheet.cell(column=_column, row=_row).value
            terminalPrint(True, "read value = {:f}".format(float(_readValue)))
            return _readValue
        except:
            terminalPrint(False, "read value")
            exit()


class _approximateFomula:  #近似式の導出
    def __init__(self): #t=0の行を指定する  返り値:なし
        self._readStartRow:int = 1251
    
    def openSheet(self, _filePath:str, _frequency:int): #対象のシートを開く 返り値:なし
        self._period:float = 1.0 / float(_frequency)
        self._excelFilePath = _filePath
        try:
            self._workBook = openpyxl.load_workbook(self._excelFilePath)
            self._mainWorkSheet = self._workBook["Sheet1"]
            terminalPrint(True, "open sheet")
        except:
            terminalPrint(False, "open sheet")
            eixt()

    def findOneCycle(self): #1周期分の行の範囲を導出    返り値:なし
        _nowColumn = self._readStartRow
        
        while(True):
            if(float(self._mainWorkSheet.cell(column=5, row=_nowColumn).value) > self._period):
                _nowColumn -= 1
                break
            else:
                _nowColumn += 1
        
        terminalPrint(True, "find range of 1 Cycle")
        
        self._readEndRow:int = _nowColumn
        self._toEndRow = None
    
    def extractOneValue(self):  #1周期分の想定データを別のファイルに抽出    返り値:なし
        try:
            self._workBook.create_sheet(title="forCalculation")
            self._calculationSheet = self._workBook["forCalculation"]
            
            terminalPrint(True, "create forCaluculation sheet")
        except:
            terminalPrint(False, "create forCaluculation sheet")
            exit()
        
        _originalNowColumn:int = 5
        _originalNowRow:int = self._readStartRow
        _toNowColumn:int = 1
        _toNowRow:int = 1
        
        while (True):
            if(_originalNowRow <= self._readEndRow):
                self._calculationSheet.cell(column=_toNowColumn, row=_toNowRow).value = self._mainWorkSheet.cell(column=_originalNowColumn, row=_originalNowRow).value
                
                _originalNowRow += 1
                _toNowRow += 1
            else:
                break
        
        _originalNowColumn += 1
        _originalNowRow = self._readStartRow
        _toNowColumn += 1
        _toNowRow = 1
            
        while (True):
            if(_originalNowRow <= self._readEndRow):
                self._calculationSheet.cell(column=_toNowColumn, row=_toNowRow).value = self._mainWorkSheet.cell(column=_originalNowColumn, row=_originalNowRow).value
                
                _originalNowRow += 1
                _toNowRow += 1
            else:
                break
        
        if((_toNowRow - 2) == (self._readEndRow - self._readStartRow)):
            terminalPrint(True, "copy extract for the relevant value")
            self._toEndRow = _toNowRow - 1
        else:
            terminalPrint(False, "copy extract For the relevant value")
            exit()
    
    def enterApproximateFomula(self):    #近似式の次数ごとの係数をセルに入力 返り値:なし
        self._calculationSheet["E2"] = "近似式"
        self._calculationSheet["E3"] = "指数"
        self._calculationSheet["F3"] = "近似式の係数"
        
        for _i in range(10, -1, -1):
            self._calculationSheet.cell(column=5, row=(14 -_i)).value = int(_i)
            
            if(_i != 0):
                self._calculationSheet.cell(column=6, row=(14 -_i)).value = "=INDEX(LINEST(B1:B{:d},A1:A{:d}^{{10,9,8,7,6,5,4,3,2,1}}),1,E{:d})".format(self._toEndRow, self._toEndRow, 14 -_i)
            else:
                self._calculationSheet.cell(column=6, row=(14 -_i)).value = "=INDEX(LINEST(B1:B{:d},A1:A{:d}^{{10,9,8,7,6,5,4,3,2,1}}),1,11)".format(self._toEndRow, self._toEndRow)
                
        terminalPrint(True, "enter approximate formula")
        
    def enterApproximateValue(self):    #近似値をセルに入力 返り値:なし
        for _i in range(1, self._toEndRow + 1, 1):
            self._calculationSheet.cell(column=3, row=_i).value = "=F4*(A{}^E4)+F5*(A{}^E5)+F6*(A{}^E6)+F7*(A{}^E7)+F8*(A{}^E8)+F9*(A{}^E9)+F10*(A{}^E10)+F11*(A{}^E11)+F12*(A{}^E12)+F13*(A{}^E13)+F14".format(_i, _i, _i, _i, _i, _i, _i, _i, _i, _i)
        
        terminalPrint(True, "enter approximate value")
    
    def enterMaximumTime(self):  #近似式の最大値をセルに入力 返り値:なし
        self._calculationSheet["E16"] = "yの最大値"
        
        self._calculationSheet["E17"] = "=MAX(C1:C{:d})".format(self._toEndRow)
        
        terminalPrint(True, "enter y-axis max value")
        
    def enterPeakToPeak(self):  #近似値のピーク-ピーク値をセルに入力 返り値:なし
        self._calculationSheet["F16"] = "ピーク-ピーク値"
        
        self._calculationSheet["F17"] = "=ABS(MAX(C1:C{:d}) - MIN(C1:C{:d}))".format(self._toEndRow, self._toEndRow)
    
    def saveApproximateFile(self):  #近似式等を入力したファイルを保存する   返り値:保存したファイルの絶対パス
        try:
            self._workBook.save("{}_TemporaryData.xlsx".format(self.excelFilePath))
            terminalPrint(True, "save approximate file")
            return "{}_TemporaryData.xlsx".format(self.excelFilePath)
        except:
            terminalPrint(False, "save approximate file")
            exit()
        
    def closeSheet(self):   #対象のシートを閉じる   返り値:なし
        try:
            self._workBook.close()
            terminalPrint(True, "close sheet")
        except:
            terminalPrint(True, "close sheet")
            

class _readEachValue:   #ファイルから各値を読み取る
    def __init__(self):
        pass
        
    def openSheet(self, _filePath:str): #対象のシートを開く 戻り値:なし
        self._excelFilePath:str = _filePath
        
        try:
            self._workBook = openpyxl.load_workbook(self._excelFilePath)
            self._workSheet = self._workBook["forCalculation"]
            
            terminalPrint(True, "open sheet")
        except:
            terminalPrint(False, "open sheet")
            exit()
    
    def findPhasePeak(self):    #ピーク値をとるときの時間を導出 戻り値:導出された時間
        cell.openSheet(self._excelFilePath, "forCalculation")
        
        _yMaxValue:float = cell.getValue("E17")
        
        _readNowRow:int = 1
        
        while(True):
            if(cell.getValue("C{:d}".format(_readNowRow)) == None):
                terminalPrint(False, "find peak value")
                self._workBook.close()
                exit()
            elif(cell.getValue("C{:d}".format(_readNowRow)) == _yMaxValue):
                terminalPrint(True, "find peak value")
                _phasePeakValue = self._workSheet.cell(column=1, row=_readNowRow).value
                print(_readNowRow)
                break
            
            _readNowRow += 1
        
        return _phasePeakValue
    
    def findPeakPeakValue(self):    #ピーク-ピーク値を導出  戻り値:ピーク-ピーク値
        return cell.getValue("F17")
    
    def closeSheet(self):   #対象のシートを閉じる   返り値:なし
        try:
            self._workBook.close()
            
            terminalPrint(True, "close sheet")
        except:
            terminalPrint(False, "close sheet")
            exit()
            

print("*** Start Ditel Easy Excel Phase Contrast Program ***");
terminalPrint(True, "version : {}".format(SOFTWARE_VERSION))

try:
    READ_DATA1:str = sys.argv[1]
    READ_DATA2:str = sys.argv[2]
    DATA_DIRECTORY_PATH:str = sys.argv[3]
    
    if((READ_DATA1 != "CH1") and (READ_DATA1 != "CH2") and (READ_DATA1 != "MTH")):
        raise ValueError(None)
    
    if((READ_DATA2 != "CH1") and (READ_DATA2 != "CH2") and (READ_DATA2 != "MTH")):
        raise ValueError(None)
    
    terminalPrint(True, "read data information")
    terminalPrint(True, "read data directory Path = {}".format(DATA_DIRECTORY_PATH))
    terminalPrint(True, "read Data1 type = {}".format(READ_DATA1))
    terminalPrint(True, "read Data2 type = {}".format(READ_DATA2))
except:
    terminalPrint(False, "read data Information")
    print("please input data directory path")
    print('Ex : python3 ./Ditel_Easy_Excel_Phase  "Data1 Type" "Data2 Type" "data Directory Path"')
    print('only "CH1", "CH2" or "MTH" can be entered for "DATA Type" and "DATA2 Type"')
    exit()
    
DATA_BASE_FILE_PAHT:str = "{}\\{}".format(os.getcwd(), "dataBase.xlsx")

cell = _cell()
dataBase = _dataBase()
approximateFomula = _approximateFomula()
readEachValue = _readEachValue()

dataBase.openSheet()
#TODO 要変更

oscilloscopeDataFilePath = csvToXlsx(oscilloscopeFilePath(DATA_DIRECTORY_PATH, dataBase.readCellValue(2, 2), READ_DATA1, "CSV"))

approximateFomula.openSheet()
approximateFomula.findOneCycle()
approximateFomula.extractOneValue()
approximateFomula.enterApproximateFomula()
approximateFomula.enterApproximateValue()
approximateFomula.enterMaximumTime()
approximateFomula.enterPeakToPeak()

approximateFilePath = approximateFomula.saveApproximateFile()

readEachValue.openSheet()

print("phase peak = {:f}".format(readEachValue.findPhasePeak()))

print("peak peak value = {:f}".format(readEachValue.findPeakPeakValue()))

readEachValue.closeSheet()

dataBase.end()